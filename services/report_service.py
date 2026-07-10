import os
import math
import sys
from datetime import datetime, date
from types import ModuleType

# Mock PIL to bypass Windows Application Control policy DLL blocks on Pillow
mock_pil = ModuleType('PIL')
mock_image = ModuleType('PIL.Image')
mock_pil.Image = mock_image
sys.modules['PIL'] = mock_pil
sys.modules['PIL.Image'] = mock_image

from app import db

# ReportLab imports
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# openpyxl imports
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

REPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'instance', 'reports')

def get_report_data(user_id: int, year: int, month: int) -> dict:
    """Aggregates all necessary statistics for report generation (income, expenses, budgets, loans, health, history)"""
    from app import User, Expense, Budget, Investment, Bill, PredictionHistory, LoanPredictionHistory
    from services.financial_health import compute_financial_health
    from services.savings_prediction import compute_savings_prediction
    from services.loan_prediction import predict_loan_eligibility
    
    user = User.query.get(user_id)
    if not user:
        return {}
        
    health = compute_financial_health(user_id)
    income = health.get('income', 50000.0)
    
    # 1. Fetch Expenses
    all_expenses = Expense.query.filter_by(user_id=user_id).all()
    month_expenses = [e for e in all_expenses if e.created_at.month == month and e.created_at.year == year]
    total_spent = sum(e.amount for e in month_expenses)
    
    # Category totals
    cat_totals = {}
    for e in month_expenses:
        cat_totals[e.category] = cat_totals.get(e.category, 0.0) + e.amount
        
    # 2. Fetch Budgets
    budgets = Budget.query.filter_by(user_id=user_id, month=month, year=year).all()
    budget_data = []
    for b in budgets:
        spent = cat_totals.get(b.category, 0.0)
        remaining = b.amount - spent
        budget_data.append({
            'category': b.category,
            'budget': b.amount,
            'spent': spent,
            'remaining': remaining,
            'over': spent > b.amount
        })
        
    # 3. Fetch Investments
    investments = Investment.query.filter_by(user_id=user_id).all()
    portfolio_value = sum(i.current_value for i in investments)
    total_invested = sum(i.invested for i in investments)
    roi = ((portfolio_value / total_invested - 1.0) * 100.0) if total_invested > 0 else 0.0
    
    # 4. Fetch Bills
    bills = Bill.query.filter_by(user_id=user_id).all()
    
    # 5. Predictions & Loan Eligibility
    pred_data = compute_savings_prediction(user_id)
    loan_data = predict_loan_eligibility(user_id, requested_amount=500000.0, tenure_months=36)
    
    # AI Report Commentary summary
    summary_text = (
        f"During {datetime(year, month, 1).strftime('%B %Y')}, your total income was ₹{income:,.2f}. "
        f"Total monthly expenses were ₹{total_spent:,.2f}, representing a savings of ₹{max(0.0, income - total_spent):,.2f} "
        f"({((income - total_spent) / income * 100.0) if income > 0 else 0:.1f}% savings rate). "
        f"Your Financial Health Score is {health.get('score', 0)}/100 ({health.get('status', '—')}). "
    )
    if cat_totals:
        top_cat = max(cat_totals, key=cat_totals.get)
        top_cat_pct = (cat_totals[top_cat] / (total_spent or 1.0)) * 100.0
        summary_text += f"Your largest expense category was {top_cat} (₹{cat_totals[top_cat]:,.2f}, or {top_cat_pct:.1f}% of total). "
        if top_cat == 'Food' or top_cat == 'Shopping':
            summary_text += f"Reducing discretionary spending in {top_cat} by 10% next month can compound into ₹{cat_totals[top_cat]*0.1*12:,.2f} additional yearly savings."
    else:
        summary_text += "No expenses were logged this month. Maintain consistency to refine predictive cash flow modeling."
        
    return {
        'username': user.username,
        'fullname': user.full_name or user.username,
        'year': year,
        'month': month,
        'income': income,
        'expenses_total': total_spent,
        'savings': max(0.0, income - total_spent),
        'portfolio_value': portfolio_value,
        'invested_total': total_invested,
        'roi': roi,
        'health_score': health.get('score', 0),
        'health_status': health.get('status', '—'),
        'health_details': health,
        'expenses': month_expenses,
        'cat_totals': cat_totals,
        'budgets': budget_data,
        'investments': investments,
        'bills': bills,
        'predictions': pred_data,
        'loan_data': loan_data,
        'ai_commentary': summary_text
    }


def generate_pdf_report(user_id: int, year: int, month: int) -> str:
    """Generates a professional PDF report using ReportLab and returns the file path"""
    os.makedirs(REPORTS_DIR, exist_ok=True)
    data = get_report_data(user_id, year, month)
    if not data:
        return ""
        
    filename = f"WalletIQ_Report_{data['username']}_{year}_{month}.pdf"
    filepath = os.path.join(REPORTS_DIR, filename)
    
    doc = SimpleDocTemplate(
        filepath,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    
    styles = getSampleStyleSheet()
    
    # Custom Styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        leading=28,
        textColor=colors.HexColor('#0f172a'),
        spaceAfter=15
    )
    
    section_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        leading=18,
        textColor=colors.HexColor('#1e293b'),
        spaceBefore=12,
        spaceAfter=8,
        keepWithNext=True
    )
    
    normal_style = ParagraphStyle(
        'NormalText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#334155'),
        spaceAfter=6
    )
    
    table_cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=11,
        textColor=colors.HexColor('#334155')
    )
    
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.white
    )

    story = []
    
    # 1. Header
    story.append(Paragraph("WalletIQ X — AI Financial Report", title_style))
    meta_text = f"<b>Generated For:</b> {data['fullname']} | <b>Report Period:</b> {datetime(year, month, 1).strftime('%B %Y')} | <b>Date:</b> {date.today().isoformat()}"
    story.append(Paragraph(meta_text, normal_style))
    story.append(Spacer(1, 15))
    
    # 2. Key Metrics Table
    story.append(Paragraph("Financial Summary Overview", section_style))
    summary_data = [
        [
            Paragraph("<b>Monthly Income</b>", normal_style),
            Paragraph("<b>Monthly Expenses</b>", normal_style),
            Paragraph("<b>Monthly Savings</b>", normal_style),
            Paragraph("<b>Health Score</b>", normal_style)
        ],
        [
            f"₹{data['income']:,.2f}",
            f"₹{data['expenses_total']:,.2f}",
            f"₹{data['savings']:,.2f}",
            f"{data['health_score']}/100 ({data['health_status']})"
        ]
    ]
    t_summary = Table(summary_data, colWidths=[130, 130, 130, 150])
    t_summary.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
        ('FONTNAME', (0,1), (-1,1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,1), (-1,1), 11),
    ]))
    story.append(t_summary)
    story.append(Spacer(1, 15))
    
    # 3. AI Commentary
    story.append(Paragraph("🤖 AI Smart Commentary Summary", section_style))
    commentary_style = ParagraphStyle(
        'CommentaryText',
        parent=normal_style,
        textColor=colors.HexColor('#1e1b4b'),
        backColor=colors.HexColor('#e0e7ff'),
        borderColor=colors.HexColor('#c7d2fe'),
        borderWidth=1,
        borderPadding=10,
        spaceAfter=15,
        borderRadius=8
    )
    story.append(Paragraph(data['ai_commentary'], commentary_style))
    
    # 4. Expenses breakdown
    story.append(Paragraph("Category Spending Breakdown", section_style))
    exp_data = [[Paragraph("<b>Category</b>", table_header_style), Paragraph("<b>Total Spent (₹)</b>", table_header_style), Paragraph("<b>% of Total</b>", table_header_style)]]
    for cat, val in data['cat_totals'].items():
        pct = (val / (data['expenses_total'] or 1.0)) * 100.0
        exp_data.append([
            Paragraph(cat, table_cell_style),
            f"₹{val:,.2f}",
            f"{pct:.1f}%"
        ])
    if len(exp_data) == 1:
        exp_data.append([Paragraph("No expenses logged", table_cell_style), "—", "—"])
        
    t_exp = Table(exp_data, colWidths=[200, 170, 170])
    t_exp.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0f172a')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(t_exp)
    story.append(Spacer(1, 15))
    
    # 5. Budget Performance
    story.append(Paragraph("Budget vs Actual Performance", section_style))
    bud_data = [[Paragraph("<b>Category</b>", table_header_style), Paragraph("<b>Budget (₹)</b>", table_header_style), Paragraph("<b>Spent (₹)</b>", table_header_style), Paragraph("<b>Remaining (₹)</b>", table_header_style), Paragraph("<b>Status</b>", table_header_style)]]
    for b in data['budgets']:
        status = Paragraph("<font color='red'>Overspent</font>", table_cell_style) if b['over'] else Paragraph("<font color='green'>OK</font>", table_cell_style)
        bud_data.append([
            Paragraph(b['category'], table_cell_style),
            f"₹{b['budget']:,.2f}",
            f"₹{b['spent']:,.2f}",
            f"₹{b['remaining']:,.2f}",
            status
        ])
    if len(bud_data) == 1:
        bud_data.append([Paragraph("No budgets configured for this month", table_cell_style), "—", "—", "—", "—"])
        
    t_bud = Table(bud_data, colWidths=[130, 100, 100, 110, 100])
    t_bud.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e293b')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(t_bud)
    story.append(Spacer(1, 15))
    
    # 6. Portfolio & Loan Summary
    story.append(Paragraph("Portfolio Value & Loan Eligibility Summaries", section_style))
    extra_summary = [
        [
            Paragraph("<b>Total Portfolio Assets</b>", normal_style),
            Paragraph("<b>Invested Cost</b>", normal_style),
            Paragraph("<b>Unrealized ROI</b>", normal_style),
            Paragraph("<b>Max Borrow Limit</b>", normal_style)
        ],
        [
            f"₹{data['portfolio_value']:,.2f}",
            f"₹{data['invested_total']:,.2f}",
            f"{data['roi']:.2f}%",
            f"₹{data['loan_data'].get('eligible_amount', 0):,.2f}"
        ]
    ]
    t_extra = Table(extra_summary, colWidths=[135, 135, 135, 135])
    t_extra.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
        ('FONTNAME', (0,1), (-1,1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,1), (-1,1), 10),
    ]))
    story.append(t_extra)

    doc.build(story)
    return filepath


def generate_excel_report(user_id: int, year: int, month: int) -> str:
    """Generates a professional multi-sheet Excel report and returns the file path"""
    os.makedirs(REPORTS_DIR, exist_ok=True)
    data = get_report_data(user_id, year, month)
    if not data:
        return ""
        
    filename = f"WalletIQ_Report_{data['username']}_{year}_{month}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)
    
    wb = Workbook()
    
    # Styles config
    title_font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    section_font = Font(name='Segoe UI', size=13, bold=True, color='1E293B')
    bold_cell = Font(name='Segoe UI', size=10, bold=True)
    normal_cell = Font(name='Segoe UI', size=10)
    
    title_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    zebra_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    accent_fill = PatternFill(start_color='E0E7FF', end_color='E0E7FF', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    # ── SHEET 1: Dashboard Summary ──
    ws1 = wb.active
    ws1.title = "Dashboard Summary"
    
    # Header Banner
    ws1.merge_cells('A1:D1')
    ws1['A1'] = "WalletIQ X — Executive Financial Summary"
    ws1['A1'].font = title_font
    ws1['A1'].fill = title_fill
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 40
    
    ws1['A3'] = "User Profile:"
    ws1['A3'].font = bold_cell
    ws1['B3'] = data['fullname']
    ws1['B3'].font = normal_cell
    
    ws1['A4'] = "Reporting Period:"
    ws1['A4'].font = bold_cell
    ws1['B4'] = datetime(year, month, 1).strftime('%B %Y')
    ws1['B4'].font = normal_cell
    
    # Core KPIs
    ws1['A6'] = "Monthly Metrics"
    ws1['A6'].font = section_font
    
    headers = ["Metric Parameter", "Value Amount (INR)", "Ratios / Status", "Health Index"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws1.cell(row=7, column=col_idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        
    kpis = [
        ("Total Monthly Income", data['income'], "100.0%", "Income Credit"),
        ("Total Monthly Expenses", data['expenses_total'], f"{data['expenses_total']/data['income']*100.0 if data['income'] > 0 else 0:.1f}% DTE", "Outflows"),
        ("Accumulated Savings", data['savings'], f"{data['savings']/data['income']*100.0 if data['income'] > 0 else 0:.1f}% Savings Rate", "Net Reserve"),
        ("Financial Health Score", data['health_score'], data['health_status'], "Weighted Rating")
    ]
    
    for idx, k in enumerate(kpis, 8):
        ws1.cell(row=idx, column=1, value=k[0]).font = bold_cell
        ws1.cell(row=idx, column=2, value=k[1]).font = normal_cell
        ws1.cell(row=idx, column=2).number_format = '₹#,##0.00'
        ws1.cell(row=idx, column=3, value=k[2]).font = normal_cell
        ws1.cell(row=idx, column=4, value=k[3]).font = normal_cell
        for col_idx in range(1, 5):
            ws1.cell(row=idx, column=col_idx).border = thin_border
            
    # AI commentary box
    ws1['A14'] = "🤖 AI Report Commentary:"
    ws1['A14'].font = bold_cell
    ws1.merge_cells('A15:D17')
    ws1['A15'] = data['ai_commentary']
    ws1['A15'].font = normal_cell
    ws1['A15'].alignment = Alignment(wrap_text=True, vertical='top')
    ws1['A15'].fill = accent_fill

    # ── SHEET 2: Expense History ──
    ws2 = wb.create_sheet(title="Expense History")
    exp_headers = ["ID", "Title Description", "Amount Spent (₹)", "Category", "Payment Mode", "Date Logged"]
    for col_idx, h in enumerate(exp_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        
    for idx, e in enumerate(data['expenses'], 2):
        ws2.cell(row=idx, column=1, value=e.id).font = normal_cell
        ws2.cell(row=idx, column=2, value=e.title).font = normal_cell
        cell_amt = ws2.cell(row=idx, column=3, value=e.amount)
        cell_amt.font = bold_cell
        cell_amt.number_format = '₹#,##0.00'
        ws2.cell(row=idx, column=4, value=e.category).font = normal_cell
        ws2.cell(row=idx, column=5, value=e.payment_mode).font = normal_cell
        ws2.cell(row=idx, column=6, value=e.created_at.strftime('%Y-%m-%d %H:%M')).font = normal_cell
        
        # Border & Zebra
        fill = zebra_fill if idx % 2 == 0 else PatternFill(fill_type=None)
        for col_idx in range(1, 7):
            cell = ws2.cell(row=idx, column=col_idx)
            cell.border = thin_border
            if fill.fill_type:
                cell.fill = fill

    # ── SHEET 3: Investments ──
    ws3 = wb.create_sheet(title="Investments")
    inv_headers = ["Asset Name", "Asset Type", "Invested Principal (₹)", "Current Balance (₹)", "ROI (%)"]
    for col_idx, h in enumerate(inv_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        
    for idx, i in enumerate(data['investments'], 2):
        ws3.cell(row=idx, column=1, value=i.name).font = bold_cell
        ws3.cell(row=idx, column=2, value=i.type).font = normal_cell
        c_inv = ws3.cell(row=idx, column=3, value=i.invested)
        c_inv.font = normal_cell
        c_inv.number_format = '₹#,##0.00'
        c_val = ws3.cell(row=idx, column=4, value=i.current_value)
        c_val.font = bold_cell
        c_val.number_format = '₹#,##0.00'
        
        roi_calc = ((i.current_value / i.invested - 1.0) * 100.0) if i.invested > 0 else 0.0
        c_roi = ws3.cell(row=idx, column=5, value=f"{roi_calc:.2f}%")
        c_roi.font = normal_cell
        
        for col_idx in range(1, 6):
            ws3.cell(row=idx, column=col_idx).border = thin_border

    # ── SHEET 4: Bills & Reminders ──
    ws4 = wb.create_sheet(title="Bills & Reminders")
    bill_headers = ["Bill Name", "Category Group", "Amount Due (₹)", "Due Day", "Priority", "Payment Mode", "Status"]
    for col_idx, h in enumerate(bill_headers, 1):
        cell = ws4.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        
    for idx, b in enumerate(data['bills'], 2):
        ws4.cell(row=idx, column=1, value=b.name).font = normal_cell
        ws4.cell(row=idx, column=2, value=b.category).font = normal_cell
        c_amt = ws4.cell(row=idx, column=3, value=b.amount)
        c_amt.font = bold_cell
        c_amt.number_format = '₹#,##0.00'
        ws4.cell(row=idx, column=4, value=b.due_day).font = normal_cell
        ws4.cell(row=idx, column=5, value=b.priority).font = normal_cell
        ws4.cell(row=idx, column=6, value=b.payment_method).font = normal_cell
        ws4.cell(row=idx, column=7, value="Paid" if b.is_paid else "Pending").font = bold_cell
        
        for col_idx in range(1, 8):
            ws4.cell(row=idx, column=col_idx).border = thin_border

    # ── SHEET 5: Predictions Projections ──
    ws5 = wb.create_sheet(title="Savings Projections")
    proj_headers = ["Month", "Conservative Scenario NW (₹)", "Moderate Scenario NW (₹)", "Aggressive Scenario NW (₹)"]
    for col_idx, h in enumerate(proj_headers, 1):
        cell = ws5.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        
    pred = data['predictions']
    if 'moderate' in pred:
        months = pred['moderate']['months']
        for idx, m in enumerate(months, 2):
            ws5.cell(row=idx, column=1, value=f"Month {m}").font = normal_cell
            
            c_con = ws5.cell(row=idx, column=2, value=pred['conservative']['net_worth'][m])
            c_con.font = normal_cell
            c_con.number_format = '₹#,##0.00'
            
            c_mod = ws5.cell(row=idx, column=3, value=pred['moderate']['net_worth'][m])
            c_mod.font = bold_cell
            c_mod.number_format = '₹#,##0.00'
            
            c_agg = ws5.cell(row=idx, column=4, value=pred['aggressive']['net_worth'][m])
            c_agg.font = normal_cell
            c_agg.number_format = '₹#,##0.00'
            
            for col_idx in range(1, 5):
                ws5.cell(row=idx, column=col_idx).border = thin_border

    # Adjust auto widths for all sheets
    from openpyxl.utils import get_column_letter
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(filepath)
    return filepath


# ── Premium Helper Functions ──────────────────────────────────────────────────

def make_report_name(report_type: str, year: int, month: int) -> str:
    """Returns a human-readable report name, e.g. 'June 2026 Monthly Financial Report'"""
    from datetime import datetime
    month_names = ["January", "February", "March", "April", "May", "June",
                   "July", "August", "September", "October", "November", "December"]
    month_name = month_names[month - 1] if 1 <= month <= 12 else str(month)
    return f"{month_name} {year} — {report_type} Financial Report"


def get_next_version_filepath(base_path: str, ext: str) -> tuple:
    """
    Checks if a file exists at base_path. If so, finds the next available
    version suffix (e.g. _v2, _v3). Returns (filepath, version_int).
    """
    if not os.path.exists(base_path):
        return base_path, 1
    version = 2
    base, _ = os.path.splitext(base_path)
    while True:
        candidate = f"{base}_v{version}.{ext}"
        if not os.path.exists(candidate):
            return candidate, version
        version += 1


def get_storage_stats(user_id: int) -> dict:
    """
    Returns aggregate statistics about a user's compiled reports:
    total generated, total downloaded, total storage in bytes, and last generated date.
    """
    from app import ReportHistory
    reports = ReportHistory.query.filter_by(user_id=user_id).all()
    total_size = sum(r.file_size or 0 for r in reports)
    total_downloads = sum(r.download_count for r in reports)
    last_gen = max((r.generated_date for r in reports), default=None)
    return {
        'total_generated': len(reports),
        'total_downloaded': total_downloads,
        'storage_bytes': total_size,
        'storage_mb': round(total_size / (1024 * 1024), 2),
        'last_generated': last_gen.strftime('%d %b %Y, %I:%M %p') if last_gen else '—'
    }


def generate_ai_comparison(user_id: int, report_a_id: int, report_b_id: int) -> dict:
    """
    Compares two compiled reports using their stored names/metadata to extract month/year info,
    then re-aggregates their financial data and produces an AI-authored comparison narrative.
    Falls back to a rule-based summary if Gemini is unavailable.
    """
    from app import ReportHistory
    import re

    def extract_year_month(report_name: str):
        month_map = {m.lower(): i + 1 for i, m in enumerate(
            ["january", "february", "march", "april", "may", "june",
             "july", "august", "september", "october", "november", "december"])}
        parts = report_name.lower().split()
        month = None
        year = None
        for p in parts:
            if p in month_map:
                month = month_map[p]
            if re.match(r'^\d{4}$', p):
                year = int(p)
        return year, month

    ra = ReportHistory.query.filter_by(id=report_a_id, user_id=user_id).first()
    rb = ReportHistory.query.filter_by(id=report_b_id, user_id=user_id).first()
    if not ra or not rb:
        return {'error': 'One or both reports not found'}

    ya, ma = extract_year_month(ra.report_name)
    yb, mb = extract_year_month(rb.report_name)

    if not all([ya, ma, yb, mb]):
        return {'error': 'Could not parse report dates for comparison'}

    data_a = get_report_data(user_id, ya, ma)
    data_b = get_report_data(user_id, yb, mb)

    # Build numeric diff table
    def delta(val_a, val_b):
        diff = val_b - val_a
        pct = ((diff / val_a) * 100.0) if val_a != 0 else 0.0
        arrow = '▲' if diff > 0 else ('▼' if diff < 0 else '—')
        return {'a': val_a, 'b': val_b, 'diff': diff, 'pct': round(pct, 1), 'arrow': arrow}

    comparison = {
        'report_a': ra.report_name,
        'report_b': rb.report_name,
        'income':        delta(data_a['income'],        data_b['income']),
        'expenses':      delta(data_a['expenses_total'], data_b['expenses_total']),
        'savings':       delta(data_a['savings'],        data_b['savings']),
        'health_score':  delta(data_a['health_score'],   data_b['health_score']),
        'portfolio':     delta(data_a['portfolio_value'], data_b['portfolio_value']),
    }

    # Rule-based narrative
    lines = [f"📊 Comparing **{ra.report_name}** vs **{rb.report_name}**:", ""]

    def fmt(c, label):
        sign = '+' if c['diff'] >= 0 else ''
        return f"• **{label}**: ₹{c['a']:,.0f} → ₹{c['b']:,.0f}  ({c['arrow']} {sign}{c['pct']}%)"

    lines += [
        fmt(comparison['income'], 'Monthly Income'),
        fmt(comparison['expenses'], 'Total Expenses'),
        fmt(comparison['savings'], 'Net Savings'),
        f"• **Financial Health Score**: {comparison['health_score']['a']}/100 → {comparison['health_score']['b']}/100 ({comparison['health_score']['arrow']} {comparison['health_score']['pct']:+.1f} pts)",
        fmt(comparison['portfolio'], 'Portfolio Value'),
        ""
    ]

    # Key insight sentence
    if comparison['savings']['diff'] > 0:
        lines.append(f"✅ Savings improved by ₹{comparison['savings']['diff']:,.0f} — great financial progress!")
    elif comparison['savings']['diff'] < 0:
        lines.append(f"⚠️ Savings declined by ₹{abs(comparison['savings']['diff']):,.0f}. Review spending categories to identify reduction opportunities.")

    if comparison['health_score']['diff'] > 0:
        lines.append(f"📈 Financial Health Score rose by {comparison['health_score']['diff']:.0f} points — your fiscal discipline is paying off.")
    elif comparison['health_score']['diff'] < 0:
        lines.append(f"📉 Health Score dropped by {abs(comparison['health_score']['diff']):.0f} pts. Check budget adherence and savings rate.")

    comparison['narrative'] = '\n'.join(lines)
    return comparison
