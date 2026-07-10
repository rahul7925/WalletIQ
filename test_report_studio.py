import unittest
import os
from app import app, db, User, ReportHistory
from services.report_service import (
    get_report_data, generate_pdf_report, generate_excel_report,
    make_report_name, get_next_version_filepath, get_storage_stats,
    generate_ai_comparison
)

class TestReportStudio(unittest.TestCase):
    def setUp(self):
        app.config['TESTING'] = True
        self.app_context = app.app_context()
        self.app_context.push()

        # Override SQL engine to SQLite memory space
        from sqlalchemy import create_engine
        self.engine = create_engine('sqlite:///:memory:')
        db._app_engines[app][None] = self.engine
        db.create_all()

        self.user = User(
            username='report_tester',
            password='pbkdf2:sha256:16$dummyhash',
            full_name='Report Tester',
            monthly_income=65000.0,
            monthly_budget=40000.0
        )
        db.session.add(self.user)
        db.session.commit()

    def tearDown(self):
        db.session.close()
        self.app_context.pop()

    # ── Core generation tests ─────────────────────────────────────────────────

    def test_report_data_aggregation(self):
        data = get_report_data(self.user.id, 2026, 7)
        self.assertEqual(data['username'], 'report_tester')
        self.assertEqual(data['income'], 65000.0)
        self.assertEqual(data['expenses_total'], 0.0)  # no expenses in test DB
        self.assertIn('total income was ₹65,000.00', data['ai_commentary'])

    def test_pdf_generation(self):
        filepath = generate_pdf_report(self.user.id, 2026, 7)
        self.assertTrue(os.path.exists(filepath))
        self.assertTrue(filepath.endswith('.pdf'))
        self.assertTrue(os.path.getsize(filepath) > 1000)
        if os.path.exists(filepath):
            os.remove(filepath)

    def test_excel_generation(self):
        filepath = generate_excel_report(self.user.id, 2026, 7)
        self.assertTrue(os.path.exists(filepath))
        self.assertTrue(filepath.endswith('.xlsx'))
        self.assertTrue(os.path.getsize(filepath) > 1000)

        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        self.assertIn('Dashboard Summary', wb.sheetnames)
        self.assertIn('Expense History', wb.sheetnames)
        self.assertIn('Investments', wb.sheetnames)
        self.assertIn('Bills & Reminders', wb.sheetnames)
        if os.path.exists(filepath):
            os.remove(filepath)

    # ── Premium utility tests ─────────────────────────────────────────────────

    def test_make_report_name(self):
        name = make_report_name('Monthly', 2026, 6)
        self.assertIn('June', name)
        self.assertIn('2026', name)
        self.assertIn('Monthly', name)

        name2 = make_report_name('Yearly', 2025, 12)
        self.assertIn('December', name2)
        self.assertIn('2025', name2)

    def test_get_next_version_filepath_new(self):
        import tempfile, os
        with tempfile.TemporaryDirectory() as tmpdir:
            base = os.path.join(tmpdir, 'report.pdf')
            fp, version = get_next_version_filepath(base, 'pdf')
            self.assertEqual(fp, base)
            self.assertEqual(version, 1)

    def test_get_next_version_filepath_existing(self):
        import tempfile, os
        with tempfile.TemporaryDirectory() as tmpdir:
            base = os.path.join(tmpdir, 'report.pdf')
            # Create the base file
            with open(base, 'w') as f: f.write('dummy')
            fp, version = get_next_version_filepath(base, 'pdf')
            self.assertEqual(version, 2)
            self.assertIn('_v2', fp)

    def test_get_storage_stats_empty(self):
        stats = get_storage_stats(self.user.id)
        self.assertEqual(stats['total_generated'], 0)
        self.assertEqual(stats['total_downloaded'], 0)
        self.assertEqual(stats['storage_bytes'], 0)
        self.assertEqual(stats['last_generated'], '—')

    def test_get_storage_stats_with_records(self):
        r1 = ReportHistory(
            user_id=self.user.id,
            report_name='January 2026 — Monthly Financial Report',
            report_type='Monthly (PDF)',
            file_name='WalletIQ_Report_report_tester_2026_1.pdf',
            file_size=51200,
            download_count=3,
            version=1
        )
        r2 = ReportHistory(
            user_id=self.user.id,
            report_name='February 2026 — Monthly Financial Report',
            report_type='Monthly (PDF)',
            file_name='WalletIQ_Report_report_tester_2026_2.pdf',
            file_size=48000,
            download_count=1,
            version=1
        )
        db.session.add_all([r1, r2])
        db.session.commit()

        stats = get_storage_stats(self.user.id)
        self.assertEqual(stats['total_generated'], 2)
        self.assertEqual(stats['total_downloaded'], 4)
        self.assertEqual(stats['storage_bytes'], 99200)
        self.assertNotEqual(stats['last_generated'], '—')

    def test_generate_ai_comparison_missing(self):
        result = generate_ai_comparison(self.user.id, 9999, 9998)
        self.assertIn('error', result)

    def test_generate_ai_comparison_valid(self):
        r1 = ReportHistory(
            user_id=self.user.id,
            report_name='January 2026 — Monthly Financial Report',
            report_type='Monthly (PDF)',
            file_name='r1.pdf',
            file_size=10000,
            version=1
        )
        r2 = ReportHistory(
            user_id=self.user.id,
            report_name='February 2026 — Monthly Financial Report',
            report_type='Monthly (PDF)',
            file_name='r2.pdf',
            file_size=10000,
            version=1
        )
        db.session.add_all([r1, r2])
        db.session.commit()

        result = generate_ai_comparison(self.user.id, r1.id, r2.id)
        self.assertNotIn('error', result)
        self.assertIn('income', result)
        self.assertIn('expenses', result)
        self.assertIn('savings', result)
        self.assertIn('narrative', result)
        self.assertIsInstance(result['narrative'], str)

if __name__ == '__main__':
    unittest.main()
